"""Spreadsheet converter for XLSX and legacy XLS workbooks."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from app.conversion.converters.text_data import rows_to_markdown_table
from app.conversion.registry import BaseConverter
from app.conversion.result import UniversalConversionResult
from app.conversion.stream_info import StreamInfo


class SpreadsheetConverter(BaseConverter):
    """Convert spreadsheet sheets to Markdown tables using format-native readers."""

    engine_name = "spreadsheet"
    priority = 10
    requires_marker_models = False
    requires_gpu = False
    _EXTENSIONS = frozenset({".xlsx", ".xls"})

    @property
    def supported_extensions(self) -> frozenset[str]:
        return self._EXTENSIONS

    def accepts(self, stream_info: StreamInfo, config: dict[str, Any]) -> bool:
        return stream_info.extension in self._EXTENSIONS

    def convert(
        self,
        filepath: str,
        config: dict[str, Any],
        device: str | None = None,
    ) -> UniversalConversionResult:
        ext = Path(filepath).suffix.lower()
        if ext == ".xls":
            return self._convert_xls(filepath, config)
        return self._convert_xlsx(filepath, config)

    def _convert_xlsx(self, filepath: str, config: dict[str, Any]) -> UniversalConversionResult:
        from openpyxl import load_workbook

        workbook = load_workbook(filepath, read_only=True, data_only=True)
        max_rows = int(config.get("spreadsheet_max_rows_per_sheet", 200))
        parts = [f"# {Path(filepath).stem}"]
        sheet_meta: list[dict[str, Any]] = []
        for sheet in workbook.worksheets:
            rows: list[list[Any]] = []
            for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if idx > max_rows:
                    break
                rows.append(list(row))
            parts.append(f"\n## Sheet: {sheet.title}\n")
            table = rows_to_markdown_table(rows)
            parts.append(table if table else "_Empty sheet._")
            truncated = bool(sheet.max_row and sheet.max_row > max_rows)
            if truncated:
                parts.append(f"\n_Only first {max_rows} rows shown._")
            sheet_meta.append(
                {
                    "name": sheet.title,
                    "rows": sheet.max_row,
                    "columns": sheet.max_column,
                    "truncated": truncated,
                }
            )
        workbook.close()
        return UniversalConversionResult(
            text="\n".join(parts).strip(),
            extension="md",
            metadata={"engine_detail": {"format": "xlsx", "sheets": sheet_meta}},
        )

    def _convert_xls(self, filepath: str, config: dict[str, Any]) -> UniversalConversionResult:
        import xlrd

        workbook = xlrd.open_workbook(filepath, on_demand=True)
        max_rows = int(config.get("spreadsheet_max_rows_per_sheet", 200))
        parts = [f"# {Path(filepath).stem}"]
        sheet_meta: list[dict[str, Any]] = []
        try:
            for sheet_name in workbook.sheet_names():
                sheet = workbook.sheet_by_name(sheet_name)
                rows: list[list[Any]] = []
                for row_idx in range(min(sheet.nrows, max_rows)):
                    rows.append(
                        [
                            _normalize_xls_cell(sheet.cell(row_idx, col_idx), workbook.datemode)
                            for col_idx in range(sheet.ncols)
                        ]
                    )
                parts.append(f"\n## Sheet: {sheet.name}\n")
                table = rows_to_markdown_table(rows)
                parts.append(table if table else "_Empty sheet._")
                truncated = sheet.nrows > max_rows
                if truncated:
                    parts.append(f"\n_Only first {max_rows} rows shown._")
                sheet_meta.append(
                    {
                        "name": sheet.name,
                        "rows": sheet.nrows,
                        "columns": sheet.ncols,
                        "truncated": truncated,
                    }
                )
        finally:
            workbook.release_resources()

        return UniversalConversionResult(
            text="\n".join(parts).strip(),
            extension="md",
            metadata={"engine_detail": {"format": "xls", "sheets": sheet_meta}},
        )


def _normalize_xls_cell(cell: Any, datemode: int) -> Any:
    import xlrd

    if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
        return None
    if cell.ctype == xlrd.XL_CELL_DATE:
        try:
            value = xlrd.xldate_as_datetime(cell.value, datemode)
        except (OverflowError, ValueError):
            return cell.value
        return value.date().isoformat() if value.time().isoformat() == "00:00:00" else value.isoformat(sep=" ")
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    if cell.ctype == xlrd.XL_CELL_ERROR:
        return xlrd.error_text_from_code.get(cell.value, f"#ERROR({cell.value})")
    if cell.ctype == xlrd.XL_CELL_NUMBER and float(cell.value).is_integer():
        return int(cell.value)
    return cell.value
